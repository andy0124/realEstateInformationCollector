import xml.etree.ElementTree as ET
from openpyxl import Workbook
import os


class parseXML2EXCEL:
    def __init__(self, xml_file_list, excel_file):
        self.xml_file_list = xml_file_list
        self.excel_file = excel_file
        # self.tree = ET.parse(self.xml_file)
        # self.root = self.tree.getroot()
        self.wb = Workbook()
        self.ws = self.wb.active

    def parse_apartment_price(self):

        xml_files = os.listdir(self.xml_file_list)
        LAWD_CD_NAME = ["종로구", "중구", "용산구", "성동구", "광진구", "동대문구", "중랑구", "성북구", "강북구", "도봉구", "노원구", "은평구", "서대문구",
                        "마포구",
                        "양천구", "강서구", "구로구", "금천구", "영등포구", "동작구", "관악구", "서초구", "강남구", "송파구", "강동구"]

        for j, subchild in enumerate(LAWD_CD_NAME):
            self.ws.cell(row=1, column=j+2).value = subchild

        row_index = 2
        yearmonth_list = list()
        for xml_file in xml_files:
            root = self.getXMLRoot(self.xml_file_list + "/" + xml_file)
            year_month = xml_file[16:22]
            if year_month not in yearmonth_list:
                yearmonth_list.append(year_month)
                row_index = row_index + 1
                self.ws.cell(row=row_index, column=1).value = year_month

            district = xml_file[23:-4]
            idx = LAWD_CD_NAME.index(district)

            price_total = 0
            total_count = 0
            for i, child in enumerate(root[1][0]):
                price_total = price_total + int(child[0].text.replace(" ", "").replace(",", ""))
                total_count = total_count + 1
                # for j, subchild in enumerate(child):
                #     self.ws.cell(row=i + 2, column=j+1).value = subchild.text
                # for j, subchild in enumerate(child):
                #     self.ws.cell(row=i + 1, column=j + 1).value = subchild.text
            self.ws.cell(row=row_index, column=idx + 2).value = price_total / total_count

        self.wb.save(self.excel_file)

    def getXMLRoot(self, xml_file):
        tree = ET.parse(xml_file)
        root = tree.getroot()
        return root

# main문 실행
if __name__ == "__main__":
    xml_file_list = "data/Apartment_Price"
    excel_file = "data.xlsx"
    parseXML2EXCEL = parseXML2EXCEL(xml_file_list, excel_file)
    parseXML2EXCEL.parse_apartment_price()